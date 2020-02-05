#!/usr/bin/env python
# -*- coding: utf-8 -*-

from datetime import date
import Main
import openpyxl
from PyQt5 import QtCore, QtWidgets
from PyQt5.QtGui import QBrush, QColor
from PyQt5.QtWidgets import QFileDialog, QTableWidgetItem, QDialog, QMessageBox
from Main import Insert_Contact, Analyze, Features


class Ui_MainWindow(object):

    # def __init__(self, button_width=100, button_height=45, table_width=1650, head_item=None):
    #     # super(Ui_MainWindow, self).__init__()
    def __init__(self):
        self.max_col = 0
        self.head_item = []
        self.vendor_list = []
        self.per_val = []
        self.vendorcontact_counts = []
        self.splitter = []
        self.msg = QMessageBox()

    def setupUi(self, MainWindow):
        MainWindow.setObjectName("Dialog")
        MainWindow.resize(645, 498)
        self.centralwidget = QtWidgets.QWidget()
        self.centralwidget.setObjectName("centralwidget")
        self.pushButton = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton.setGeometry(QtCore.QRect(0, 0, Main.button_width, Main.button_height))
        self.pushButton.setObjectName("pushButton")
        self.pushButton_2 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_2.setGeometry(QtCore.QRect(110, 0, Main.button_width, Main.button_height))
        self.pushButton_2.setObjectName("pushButton_2")
        self.pushButton_3 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_3.setGeometry(QtCore.QRect(220, 0, Main.button_width, Main.button_height))
        self.pushButton_3.setObjectName("pushButton_3")
        self.tableWidget = QtWidgets.QTableWidget(self.centralwidget)
        self.tableWidget.setEnabled(True)
        self.tableWidget.setGeometry(QtCore.QRect(0, 50, Main.table_width, 850))
        self.tableWidget.setObjectName("tableWidget")
        self.tableWidget.setColumnCount(0)
        self.tableWidget.setRowCount(0)

        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 645, 22))
        self.menubar.setObjectName("menubar")
        self.menuFile = QtWidgets.QMenu(self.menubar)
        self.menuFile.setObjectName("menuFile")
        self.menuEdit = QtWidgets.QMenu(self.menubar)
        self.menuEdit.setObjectName("menuEdit")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        self.actionOpen = QtWidgets.QAction(MainWindow)
        self.actionOpen.setObjectName("actionOpen")
        self.actionImport = QtWidgets.QAction(MainWindow)
        self.actionImport.setObjectName("actionImport")
        self.actionSave = QtWidgets.QAction(MainWindow)
        self.actionSave.setObjectName("actionSave")
        self.actionExit = QtWidgets.QAction(MainWindow)
        self.actionExit.setObjectName("actionExit")
        self.actionzengjia_waibaoshang = QtWidgets.QAction(MainWindow)
        self.actionzengjia_waibaoshang.setObjectName("actionzengjia_waibaoshang")
        self.action_insert_contact = QtWidgets.QAction(MainWindow)
        self.action_insert_contact.setObjectName("action_insert_contact")
        self.action_del_contact = QtWidgets.QAction(MainWindow)
        self.action_del_contact.setObjectName("action_del_contact")
        self.menuFile.addAction(self.actionOpen)
        self.menuFile.addAction(self.actionImport)
        self.menuFile.addAction(self.actionSave)
        self.menuFile.addAction(self.actionExit)
        self.menuEdit.addSeparator()
        self.menuEdit.addAction(self.actionzengjia_waibaoshang)
        self.menuEdit.addAction(self.action_insert_contact)
        self.menuEdit.addAction(self.action_del_contact)
        self.menubar.addAction(self.menuFile.menuAction())
        self.menubar.addAction(self.menuEdit.menuAction())
        self.actionOpen.triggered.connect(self.openfile)
        self.action_insert_contact.triggered.connect(self.add_contact)  # connect
        self.action_del_contact.triggered.connect(self.del_contact)
        self.actionSave.triggered.connect(self.save_file)
        self.pushButton.clicked.connect(self.add_contact)
        self.pushButton_2.clicked.connect(self.del_contact)
        self.pushButton_3.clicked.connect(self.open_analyze)
        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "供应商管理系统"))
        self.pushButton.setText(_translate("MainWindow", "增加合同"))
        self.pushButton_2.setText(_translate("MainWindow", "删除合同"))
        self.pushButton_3.setText(_translate("MainWindow", "数据中心"))
        self.menuFile.setTitle(_translate("MainWindow", "File"))
        self.menuEdit.setTitle(_translate("MainWindow", "Edit"))
        self.actionOpen.setText(_translate("MainWindow", "Open"))
        self.actionImport.setText(_translate("MainWindow", "Import"))
        self.actionSave.setText(_translate("MainWindow", "Save"))
        self.actionExit.setText(_translate("MainWindow", "Exit"))
        self.actionzengjia_waibaoshang.setText(_translate("MainWindow", "zengjia waibaoshang"))
        self.action_insert_contact.setText(_translate("MainWindow", "增加合同"))
        self.action_del_contact.setText(_translate("MainWindow", "删除合同"))

    def openfile(self):  # import vendor excel, Please don't
        excel_file, _ = QFileDialog.getOpenFileName(None,
                                                    'Choose Files',
                                                    'E:\\Test\\',
                                                    'Excel Files (*.xlsx)')
        wb = openpyxl.load_workbook(filename=excel_file)
        ws = wb.active
        self.max_col = ws.max_column
        self.tableWidget.setRowCount(ws.max_row)
        self.tableWidget.setColumnCount(self.max_col)
        for row in ws.iter_rows(max_row=1):
            for cell in row:
                self.head_item.append(cell.value)
        self.tableWidget.setHorizontalHeaderLabels(self.head_item)  # get excel head item
        for i in range(2, ws.max_row + 1):
            for x in range(1, self.max_col + 1):
                if ws.cell(i, x).value is None:
                    self.tableWidget.setItem(i - 2, x - 1, QTableWidgetItem(' '))
                elif isinstance(ws.cell(i, x).value, date):
                    ymd = date.isoformat(ws.cell(i, x).value)
                    self.tableWidget.setItem(i - 2, x - 1, QTableWidgetItem(ymd))
                else:
                    self.tableWidget.setItem(i - 2, x - 1, QTableWidgetItem(str(ws.cell(i, x).value)))

        # below is for calculate per day price
        for col in range(self.tableWidget.columnCount()):
            if self.tableWidget.horizontalHeaderItem(col).text() == '总人天':
                self.per_val.append(col)
            if self.tableWidget.horizontalHeaderItem(col).text() == '个数':
                self.per_val.append(col)
            if self.tableWidget.horizontalHeaderItem(col).text() == '平均人天':
                self.per_val.append(col)
        try:  # could not convert space to float
            for row in range(self.tableWidget.rowCount()):
                a = self.tableWidget.item(row, self.per_val[0]).text()
                b = self.tableWidget.item(row, self.per_val[1]).text()
                if a != ' ' and b != ' ':
                    result = float(a) / float(b)
                    result = round(result, 1)  # 2.53333 to 2.5
                    result = str(result)
                    self.tableWidget.setItem(row, self.per_val[2], QTableWidgetItem(result))
        except Exception as e:
            print(e)

        for merged_cell in ws.merged_cells:
            r1, r2, c1, c2 = merged_cell.min_row, merged_cell.max_row, merged_cell.min_col, merged_cell.max_col
            # print(r1, r2, c1, c2)
            self.vendorcontact_counts.append(r2 - r1)
            self.tableWidget.setSpan(r1 - 2, c1 - 1, r2 - (r1 - 1), c2)
            self.vendor_list.append(self.tableWidget.item(r1 - 2, c1 - 1).text())
            self.splitter.append(r2)  # every vendor last row
        for row in self.splitter:  # fill background for null space
            for i in range(0, self.max_col):
                item = QTableWidgetItem()
                item.setBackground(QBrush(QColor(128, 128, 128)))
                self.tableWidget.setItem(row - 1, i, item)
        self.tableWidget.resizeColumnsToContents()  # cell width follow the content length

    def import_file(self):
        pass

    def save_file(self):
        path = QFileDialog.getSaveFileName(None,
                                           'SaveFile',
                                           'E:\\Test\\',
                                           'Excel Files (*.xlsx)')
        try:
            wb = openpyxl.Workbook()
            ws = wb.create_sheet(index=0, title="VendorSystem")
            for head_val in range(self.tableWidget.columnCount()):
                head = self.tableWidget.horizontalHeaderItem(head_val).text()
                ws.cell(1, head_val + 1, head)
            for row in range(self.tableWidget.rowCount()):
                for column in range(self.tableWidget.columnCount()):
                    item = self.tableWidget.item(row, column)
                    if item is not None:
                        ws.cell(row + 2, column + 1, item.text())
                    else:
                        ws.cell(row + 2, column, ' ')
            for i in range(len(self.splitter)):  # merge cell column 1
                first = self.splitter[i] - self.vendorcontact_counts[i]
                ws.merge_cells('A%d:A%d' % (first, self.splitter[i]))
            wb.save(path[0])
        except Exception as e:
            print(e)

    def exit_file(self):
        pass

    def open_analyze(self):
        # url_string = "file:///C:/Users/Administrator/PycharmProjects/Vender_System/Main/render.html"
        # a = Analyze.BrowserWindow(url_string)
        # a.exec()
        sum_list = []
        manday_list = []
        vendor_dict = self.get_vendor_info()
        features = Features.Features()
        for i in vendor_dict:
            _sum = features.get_sum(vendor_dict[i]['付款金额'])
            _manday = features.get_sum(vendor_dict[i]['总人天'])
            sum_list.append(_sum)
            manday_list.append(_manday)
        a = Analyze.BrowserWindow(self.vendor_list, sum_list, manday_list)
        a.exec()
        # print(self.vendor_dict)
        # for vendor_name in self.vendor_list:
        #     _list = self.get_onevendor_column_info()
        #     _list = list(map(float, _list))
        #     sum_list.append(sum(_list))
        # print(sum_list)
        # day_list = []
        # for vendor_name in self.vendor_list:
        #     _list = self.get_onevendor_column_info(vendor_name, '总人天')
        #     _list = list(map(float, _list))
        #     day_list.append(sum(_list))
        #
        # frame = QtWidgets.QWidget()
        # self.analyze = new.Ui_Form(self.vendor_list, sum_list, day_list)
        # self.analyze.setupUi(frame)
        # frame.show()

    def add_contact(self):
        frame = QDialog()
        self.insert_dialog = Insert_Contact.Ui_Dialog(self.head_item, self.max_col, self.vendor_list)
        self.insert_dialog.setupUi(frame)
        self.insert_dialog.buttonBox.accepted.connect(self.add_contact_ok)
        frame.exec()

    def add_contact_ok(self):
        try:
            vendor_val = self.insert_dialog.comboBox.currentIndex()  # Insert_Contact combo box item
            insert_row = self.splitter[vendor_val] - 1
            firstpostion = insert_row - self.vendorcontact_counts[vendor_val] - 1
            self.tableWidget.insertRow(insert_row)
            self.vendorcontact_counts[vendor_val] += 1
            for i in range(vendor_val, len(self.splitter)):  # every splitter need add 1
                self.splitter[i] += 1
            # print(firstpostion, 0, self.vendorcontact_counts[vendor_val] + 1, 1)
            self.tableWidget.setSpan(firstpostion, 0, self.vendorcontact_counts[vendor_val] + 1, 1)
            for val in range(1, self.insert_dialog.tableWidget.columnCount()):
                new_content = self.insert_dialog.tableWidget.item(0, val).text()  # what you type in dialog
                self.tableWidget.setItem(insert_row, val, QTableWidgetItem(new_content))
                # print(new_content)
        except Exception as e:
            self.msg.warning(self.msg, '提示', '你有空白未填写', self.msg.Ok)
            print(e)

    def add_vendor(self):
        pass

    def del_contact(self):
        only_val = []
        for col in range(self.tableWidget.columnCount()):
            if self.tableWidget.horizontalHeaderItem(col).text() == '凭证字号':
                only_val.append(col)
        if self.tableWidget.currentColumn() != only_val[0]:
            self.msg.warning(self.msg, '提示', "'请选择一个'凭证字号'的单元格再删除此行", QMessageBox.Ok)
        else:
            fianl_warning = self.msg.warning(self.msg, '提示', "确定删除此行？", self.msg.Yes | self.msg.No)
            if fianl_warning == self.msg.Yes:
                selected_row = self.tableWidget.currentRow()
                self.tableWidget.removeRow(selected_row)
            elif fianl_warning == self.msg.No:
                self.msg.close()

    def get_vendor_info(self):  # {公司名：{表头:[]}}
        first_end = []
        vendor_dict = {}
        for i in range(len(self.splitter)):
            first = self.splitter[i] - self.vendorcontact_counts[i] - 2
            end = self.splitter[i] - 1
            first_end.append([first, end])
        print(first_end)  # '上海大宁文化传播有限公司'     first_end[0]= [0, 88]

        temp_list = []
        for x in range(len(self.vendor_list)):
            head_list = {}
            for col in range(1, self.tableWidget.columnCount()):
                head_name = self.tableWidget.horizontalHeaderItem(col).text()
                _list = []
                for row in range(first_end[x][0], first_end[x][1]):
                    item = self.tableWidget.item(row, col).text()
                    _list.append(item)
                head_list.setdefault(head_name, _list)
            temp_list.append(head_list)

        for x in range(len(self.vendor_list)):
            vendor_dict.setdefault(self.vendor_list[x], temp_list[x])
        return vendor_dict
        # print(self.vendor_dict['漕河三维动画有限公司']['凭证字号'])
